import json
import tkinter as tk
from tkinter import ttk
from ttkthemes import ThemedTk
from tkinter import filedialog
from datetime import datetime, timedelta
import calendar
from tkcalendar import Calendar  # 使用 tkcalendar 替代 ttk.Calendar
import pandas as pd
from openpyxl import Workbook
from tkinter import messagebox  # 引入 messagebox 模块
import webbrowser  # 引入 webbrowser 模块

# 创建主窗口
root = ThemedTk(theme="sun-valley")
root.title("JSON 数据查看器")
root.geometry("800x600")

# 设置窗口图标
root.iconphoto(False, tk.PhotoImage(file='logo.png'))

# 居中窗口函数
def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")

# 设置主窗口居中
center_window(root, 800, 800)

# 获取当前周的日期范围
def get_current_week():
    today = datetime.today()
    start_of_week = today - timedelta(days=today.weekday())  # 本周周一
    end_of_week = start_of_week + timedelta(days=6)  # 本周周日
    return start_of_week, end_of_week

# 更新日期标签
def update_week_label(start_date, end_date):
    week_label.config(text=f"选择的周: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")

# 日期选择功能
def select_week():
    def on_date_selected():
        selected_date = calendar.selection_get()
        start_of_week = selected_date - timedelta(days=selected_date.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        update_week_label(start_of_week, end_of_week)
        calendar_window.destroy()

    calendar_window = tk.Toplevel(root)
    calendar_window.title("选择日期")
    calendar_window.resizable(False, False)
    center_window(calendar_window, 400, 300)

    calendar = Calendar(calendar_window, selectmode="day", firstweekday="monday")
    calendar.pack(pady=10)

    ttk.Button(calendar_window, text="确定", command=on_date_selected).pack(pady=10)

# 搜索功能
def search_data():
    query = search_var.get().strip()
    if not query:
        status_label.config(text="请输入搜索内容！", foreground="red")
        return

    for row in table.get_children():
        table.delete(row)

    try:
        for key, record in current_data["_default"].items():
            if query in str(record["姓名"]) or query in str(record["学号"]):
                table.insert("", "end", values=(
                    record["姓名"],
                    record["学号"],
                    record["免白天岗状态"],
                    record["免夜间岗状态"],
                    record["站岗状态"],
                    record["站白天岗次数"],
                    record["站夜间岗次数"]
                ))
        status_label.config(text="搜索完成！", foreground="green")
    except Exception as e:
        status_label.config(text=f"搜索失败: {str(e)}", foreground="red")

# 刷新功能
def refresh_table():
    try:
        update_table(current_data)
        status_label.config(text="表格已刷新！", foreground="green")
    except Exception as e:
        status_label.config(text=f"刷新失败: {str(e)}", foreground="red")

# 加载 JSON 文件数据
def load_json():
    global current_data
    file_path = filedialog.askopenfilename(
        filetypes=[("JSON 文件", "*.json")],
        title="选择 JSON 文件"
    )
    if not file_path:
        status_label.config(text="加载取消！", foreground="red")
        return

    try:
        with open(file_path, "r", encoding="utf-8") as file:
            current_data = json.load(file)
            update_table(current_data)
            status_label.config(text="JSON 数据加载成功！", foreground="green")
    except Exception as e:
        status_label.config(text=f"加载失败: {str(e)}", foreground="red")

# 加载默认 JSON 文件
def load_default_json():
    global current_data
    try:
        with open("database.json", "r", encoding="utf-8") as file:
            current_data = json.load(file)
            if "_default" not in current_data or not isinstance(current_data["_default"], dict):
                raise ValueError("JSON 文件格式错误")
            update_table(current_data)
            status_label.config(text="默认 JSON 数据加载成功！", foreground="green")
    except FileNotFoundError:
        messagebox.showerror("错误", "默认 JSON 文件未找到！")
        status_label.config(text="默认 JSON 文件未找到！", foreground="red")
    except (ValueError, json.JSONDecodeError) as e:
        messagebox.showerror("错误", f"加载默认 JSON 文件失败：{str(e)}")
        status_label.config(text="加载默认 JSON 文件失败！", foreground="red")
    except Exception as e:
        messagebox.showerror("错误", f"发生未知错误：{str(e)}")
        status_label.config(text="加载默认 JSON 文件失败！", foreground="red")

# 更新表格数据
def update_table(data):
    for row in table.get_children():
        table.delete(row)

    for key, record in data["_default"].items():
        table.insert("", "end", values=(
            record["姓名"],
            record["学号"],
            record["免白天岗状态"],
            record["免夜间岗状态"],
            record["站岗状态"],
            record["站白天岗次数"],
            record["站夜间岗次数"]
        ))

# 输出岗表功能
def export_duty_schedule():
    start_of_week, end_of_week = get_current_week()
    default_filename = f"{start_of_week.strftime('%Y-%m-%d')}_至_{end_of_week.strftime('%Y-%m-%d')}_岗表.xlsx"
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel 文件", "*.xlsx")],
        title="选择输出路径",
        initialfile=default_filename
    )
    if not file_path:
        status_label.config(text="输出取消！", foreground="red")
        return

    try:
        # 加载数据库
        with open("database.json", "r", encoding="utf-8") as file:
            database = json.load(file)

        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "岗表"

        # 写入当周日期到 A4-A10，仅包含“XX日”
        current_date = start_of_week
        for row in range(4, 11):  # A4 到 A10
            ws[f"A{row}"] = current_date.strftime("%d日")
            current_date += timedelta(days=1)

        # 写入固定时间段到 B3-H3
        time_slots = ["0-2", "2-4", "4-6", "6-13", "13-19", "22-24"]
        for col, slot in enumerate(time_slots, start=2):  # B 列到 H 列
            ws.cell(row=3, column=col, value=slot)

        # 按顺序写入姓名到 B4-D10 和 G4-G10，先横向后纵向
        rows = range(4, 11)
        cols = ["B", "C", "D", "G"]
        cells = [f"{col}{row}" for row in rows for col in cols]  # 先横向后纵向生成单元格列表

        # 获取数据库记录并排序
        records = list(database["_default"].values())
        records.sort(key=lambda x: x["站夜间岗次数"])  # 按站夜间岗次数升序排序

        cell_index = 0
        while cell_index < len(cells):  # 循环直到所有单元格写满
            for record in records:
                # 跳过站岗状态为“否”的记录
                if record["站岗状态"] == "否":
                    record["站夜间岗次数"] += 1
                    continue

                # 跳过免夜间岗状态不为 0 的记录
                if record["免夜间岗状态"] != 0:
                    record["站夜间岗次数"] += 1
                    record["免夜间岗状态"] -= 1
                    continue

                # 写入姓名到单元格
                if cell_index < len(cells):
                    ws[cells[cell_index]] = record["姓名"]
                    cell_index += 1
                    # 更新站夜间岗次数
                    record["站夜间岗次数"] += 1
                else:
                    break  # 如果单元格写满，停止写入

        # 按顺序写入姓名到 E4-F10，先横向后纵向
        rows = range(4, 11)
        cols = ["E", "F"]
        cells = [f"{col}{row}" for row in rows for col in cols]  # 先横向后纵向生成单元格列表

        # 获取数据库记录并排序
        records = list(database["_default"].values())
        records.sort(key=lambda x: x["站白天岗次数"])  # 按站白天岗次数升序排序

        cell_index = 0
        while cell_index < len(cells):  # 循环直到所有单元格写满
            for record in records:
                # 跳过站岗状态为“否”的记录
                if record["站岗状态"] == "否":
                    record["站白天岗次数"] += 1
                    continue

                # 跳过免白天岗状态不为 0 的记录
                if record["免白天岗状态"] != 0:
                    record["站白天岗次数"] += 1
                    record["免白天岗状态"] -= 1
                    continue

                # 写入姓名到单元格
                if cell_index < len(cells):
                    ws[cells[cell_index]] = record["姓名"]
                    cell_index += 1
                    # 更新站白天岗次数
                    record["站白天岗次数"] += 1
                else:
                    break  # 如果单元格写满，停止写入

        # 保存更新后的数据库
        with open("database.json", "w", encoding="utf-8") as file:
            json.dump(database, file, ensure_ascii=False, indent=4)

        # 保存 Excel 文件
        wb.save(file_path)

        # 刷新 Treeview 数据
        update_table(database)

        status_label.config(text="岗表成功输出！", foreground="green")
    except Exception as e:
        status_label.config(text=f"输出失败: {str(e)}", foreground="red")

# 更新免白天岗状态
def increment_exempt_day():
    name = name_var.get().strip()
    if not name:
        messagebox.showerror("错误", "请输入有效的姓名！")
        return

    found = False
    for key, record in current_data["_default"].items():
        if record["姓名"] == name:
            record["免白天岗状态"] += 1
            found = True
            break

    if found:
        # 同步到 JSON 文件
        with open("database.json", "w", encoding="utf-8") as file:
            json.dump(current_data, file, ensure_ascii=False, indent=4)
        update_table(current_data)
        status_label.config(text=f"{name} 的免白天岗状态已更新！", foreground="green")
    else:
        messagebox.showerror("错误", "没有找到对应数据！")

# 更新免夜间岗状态
def increment_exempt_night():
    name = name_var.get().strip()
    if not name:
        messagebox.showerror("错误", "请输入有效的姓名！")
        return

    found = False
    for key, record in current_data["_default"].items():
        if record["姓名"] == name:
            record["免夜间岗状态"] += 1
            found = True
            break

    if found:
        # 同步到 JSON 文件
        with open("database.json", "w", encoding="utf-8") as file:
            json.dump(current_data, file, ensure_ascii=False, indent=4)
        update_table(current_data)
        status_label.config(text=f"{name} 的免夜间岗状态已更新！", foreground="green")
    else:
        messagebox.showerror("错误", "没有找到对应数据！")

# 当选中 Treeview 中的数据时，将姓名填入输入框
def on_treeview_select(event):
    selected_item = table.selection()
    if selected_item:
        item = table.item(selected_item[0])
        name_var.set(item["values"][0])  # 将选中数据的姓名填入输入框

# 关于窗口函数
def show_about():
    about_window = tk.Toplevel(root)
    about_window.title("关于")
    about_window.resizable(False, False)

    # 设置窗口大小并居中
    window_width, window_height = 400, 200
    center_window(about_window, window_width, window_height)

    # 显示文本
    tk.Label(about_window, text="JSON 数据查看器 v1.0", font=("Arial", 14)).pack(pady=10)
    tk.Label(about_window, text="作者：埃及猪肉", font=("Arial", 12)).pack(pady=5)
    tk.Label(about_window, text="有什么 Bug 可以到作者的博客反馈哦", font=("Arial", 12)).pack(pady=5)

    # 超链接
    def open_url():
        webbrowser.open("https://pattianfang.github.io/")

    link = tk.Label(about_window, text="https://pattianfang.github.io/", font=("Arial", 12), fg="blue", cursor="hand2")
    link.pack(pady=5)
    link.bind("<Button-1>", lambda e: open_url())

    # 确定按钮
    ttk.Button(about_window, text="确定", command=about_window.destroy).pack(pady=10)

# 界面布局
main_frame = ttk.Frame(root, padding="10")
main_frame.pack(fill="both", expand=True)

# 数据预览区域
table_frame = ttk.Frame(main_frame)
table_frame.pack(fill="both", expand=True, pady=10)

columns = ("姓名", "学号", "免白天岗状态", "免夜间岗状态", "站岗状态", "站白天岗次数", "站夜间岗次数")
table = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)

# 添加滑动条
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side="right", fill="y")

for col in columns:
    table.heading(col, text=col)
    table.column(col, anchor="center", width=100)

table.pack(fill="both", expand=True)

# 搜索框和按钮区域
search_frame = ttk.Frame(main_frame)
search_frame.pack(fill="x", pady=10)

search_var = tk.StringVar()
ttk.Entry(search_frame, textvariable=search_var, width=30).pack(side="left", padx=5)
ttk.Button(search_frame, text="搜索", command=search_data).pack(side="left", padx=5)
ttk.Button(search_frame, text="刷新", command=refresh_table).pack(side="left", padx=5)  # 添加刷新按钮

# 主界面日期选择区域
date_frame = ttk.Frame(main_frame)
date_frame.pack(fill="x", pady=10)

start_of_week, end_of_week = get_current_week()
week_label = ttk.Label(date_frame, text="选择的周: ")
week_label.pack(side="left", padx=10)
update_week_label(start_of_week, end_of_week)

ttk.Button(date_frame, text="选择日期", command=select_week).pack(side="left", padx=10)

# 新增免岗操作区域
exempt_frame = ttk.Frame(main_frame)
exempt_frame.pack(fill="x", pady=10)

name_var = tk.StringVar()
ttk.Entry(exempt_frame, textvariable=name_var, width=30).pack(side="left", padx=5)
ttk.Button(exempt_frame, text="免白岗", command=increment_exempt_day).pack(side="left", padx=5)
ttk.Button(exempt_frame, text="免夜岗", command=increment_exempt_night).pack(side="left", padx=5)

# 新增输入值班员区域
duty_frame = ttk.Frame(main_frame)
duty_frame.pack(fill="x", pady=10)

ttk.Label(duty_frame, text="输入值班员:").pack(side="left", padx=5)
duty_var = tk.StringVar()
ttk.Entry(duty_frame, textvariable=duty_var, width=30).pack(side="left", padx=5)

# 绑定 Treeview 选择事件
table.bind("<<TreeviewSelect>>", on_treeview_select)

# 按钮区域
button_frame = ttk.Frame(main_frame)
button_frame.pack(fill="x", pady=10)

ttk.Button(button_frame, text="加载 JSON 文件", command=load_json).pack(side="left", padx=10)
ttk.Button(button_frame, text="输出岗表", command=export_duty_schedule).pack(side="left", padx=10)

# 状态区域
status_frame = ttk.Frame(main_frame)
status_frame.pack(fill="x", pady=10)

status_label = ttk.Label(status_frame, text="")
status_label.pack()

# 关于按钮
about_button = ttk.Button(root, text="关于", command=show_about)
about_button.pack(side="bottom", pady=10)

# 启动主循环前加载默认 JSON 文件
load_default_json()
root.mainloop()
